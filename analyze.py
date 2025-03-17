import json
import pandas as pd
import os
import math
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Function to parse JSON results file
def parse_results(filepath):
    try:
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        protocol = data.get('protocol', 'Unknown')
        file_prefix = data.get('file_prefix', 'Unknown')
        
        results = {}
        
        for file_size, file_data in data.get('files', {}).items():
            results[file_size] = {
                'transfer_time': file_data['transfer_time'],
                'throughput': file_data['throughput_bps'],
                'overhead_ratio': file_data['overhead_ratio']
            }
        
        return results, protocol, file_prefix
    except Exception as e:
        print(f"Error parsing {filepath}: {e}")
        return {}, "Unknown", "Unknown"

# Function to merge data from multiple result files for the same protocol
def merge_protocol_results(file_results):
    """Merge results from different files (A and B) for the same protocol"""
    if not file_results:
        return {}
    
    # Start with the first file's results
    merged_data = {}
    
    # Get all unique file sizes from all results
    all_file_sizes = set()
    for result in file_results:
        all_file_sizes.update(result.keys())
    
    # Process each file size
    for file_name in all_file_sizes:
        # Extract the size part (e.g., "10kB") by removing the prefix
        if '_' in file_name:
            base_size = file_name.split('_')[1]
        else:
            base_size = file_name
        
        base_size_conversion = {
            '10240': '10kB',
            '102400': '100kB',
            '1048576': '1MB',
            '10485760': '10MB'
        }

        # Skip if base_size is not in the expected format
        if base_size not in ['10kB', '100kB', '1MB', '10MB']:
            base_size = base_size_conversion[base_size]
        
        if base_size not in ['10kB', '100kB', '1MB', '10MB']:
            continue
            
        # Find all results for this file size across different result files
        size_results = []
        for result in file_results:
            # Look for exact match or any match with this base size
            if file_name in result:
                size_results.append(result[file_name])
            else:
                # Try to find any key that contains this base size
                for key in result:
                    if base_size in key:
                        size_results.append(result[key])
                        break
        
        if not size_results:
            continue
        
        # Initialize data for this file size
        merged_data[base_size] = {}
        
        # Merge metrics
        for metric in ['transfer_time', 'throughput', 'overhead_ratio']:
            # Get all available values for this metric
            means = [r[metric]['mean'] for r in size_results if r[metric]['mean'] != 0]
            stddevs = [r[metric]['stddev'] for r in size_results if r[metric]['stddev'] != 0]
            
            if means:
                # Average of means
                combined_mean = sum(means) / len(means)
                
                # Combined standard deviation
                if stddevs and any(s > 0 for s in stddevs):
                    # Using propagation of uncertainty formula for combined measurements
                    combined_stddev = math.sqrt(sum(s**2 for s in stddevs) / len(stddevs))
                else:
                    combined_stddev = 0
                
                merged_data[base_size][metric] = {
                    'mean': combined_mean,
                    'stddev': combined_stddev
                }
            else:
                merged_data[base_size][metric] = {'mean': 0, 'stddev': 0}
    
    return merged_data

# Function to create excel from the data
def create_excel(protocol_data, output_filename="results.xlsx"):
    # Get all available protocols
    protocols = list(protocol_data.keys())
    
    # Create a multiindex for columns
    column_tuples = []
    for protocol in protocols:
        column_tuples.extend([(protocol, 'Mean'), (protocol, 'Std Dev')])
    
    columns = pd.MultiIndex.from_tuples(column_tuples)
    
    # Get all unique file sizes
    file_sizes = set()
    for protocol in protocols:
        file_sizes.update(protocol_data[protocol].keys())
    
    # Sort file sizes in logical order
    size_order = {"10kB": 0, "100kB": 1, "1MB": 2, "10MB": 3}
    file_sizes = sorted(list(file_sizes), key=lambda x: size_order.get(x, 99))
    
    # Prepare data for each metric
    transfer_time_data = []
    throughput_data = []
    overhead_data = []
    
    for file_size in file_sizes:
        # Transfer time row
        row = []
        for protocol in protocols:
            if file_size in protocol_data[protocol]:
                row.extend([
                    protocol_data[protocol][file_size]['transfer_time']['mean'],
                    protocol_data[protocol][file_size]['transfer_time']['stddev']
                ])
            else:
                row.extend([None, None])
        transfer_time_data.append(row)
        
        # Throughput row
        row = []
        for protocol in protocols:
            if file_size in protocol_data[protocol]:
                row.extend([
                    protocol_data[protocol][file_size]['throughput']['mean'],
                    protocol_data[protocol][file_size]['throughput']['stddev']
                ])
            else:
                row.extend([None, None])
        throughput_data.append(row)
        
        # Overhead ratio row
        row = []
        for protocol in protocols:
            if file_size in protocol_data[protocol]:
                row.extend([
                    protocol_data[protocol][file_size]['overhead_ratio']['mean'],
                    protocol_data[protocol][file_size]['overhead_ratio']['stddev']
                ])
            else:
                row.extend([None, None])
        overhead_data.append(row)
    
    # Create DataFrames
    transfer_time_df = pd.DataFrame(transfer_time_data, index=file_sizes, columns=columns)
    throughput_df = pd.DataFrame(throughput_data, index=file_sizes, columns=columns)
    overhead_df = pd.DataFrame(overhead_data, index=file_sizes, columns=columns)
    
    # Create Excel file
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Write each dataset to a different sheet
        transfer_time_df.to_excel(writer, sheet_name='Transfer Time (s)')
        throughput_df.to_excel(writer, sheet_name='Throughput (bps)')
        overhead_df.to_excel(writer, sheet_name='Overhead Ratio')
        
        # Get the workbook
        workbook = writer.book
        
        # Style each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Format headers
            for col in range(1, len(columns) + 2):  # +2 for index and file size columns
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format index column
            for row in range(2, len(file_sizes) + 2):
                cell = sheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format all data cells
            for row in range(2, len(file_sizes) + 2):
                for col in range(2, len(columns) + 2):
                    cell = sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Format numbers according to sheet type
                    if sheet_name == 'Transfer Time (s)':
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00000'
                    elif sheet_name == 'Throughput (bps)':
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    elif sheet_name == 'Overhead Ratio':
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00000'
            
            # Set column widths
            for col in range(1, len(columns) + 2):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 15
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(file_sizes) + 2):
                for col in range(1, len(columns) + 2):
                    sheet.cell(row=row, column=col).border = thin_border
    
    print(f"Excel file created: {output_filename}")

def main():
    # Find all result JSON files
    result_files = glob.glob("**/results_*_from_*_*.json")
    
    print(f"Found {len(result_files)} result files:")
    for file in result_files:
        print(f" - {file}")
    
    # Group files by protocol
    protocol_files = {}
    
    for file in result_files:
        file_data, protocol, prefix = parse_results(file)
        
        if protocol not in protocol_files:
            protocol_files[protocol] = []
        
        protocol_files[protocol].append(file_data)
    from pprint import pprint
    pprint(protocol_files)
    
    # Merge results for each protocol
    protocol_data = {}
    
    for protocol, file_results in protocol_files.items():
        print(f"Merging {len(file_results)} files for {protocol}...")
        protocol_data[protocol] = merge_protocol_results(file_results)
    
    # Create Excel file
    create_excel(protocol_data)

main()